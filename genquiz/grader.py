from openpyxl import load_workbook
import numpy as np
import os
import pickle
from .generator import *

percent_acc = 0.07
def is_numeric(obj):
    attrs = [np.int, np.int32, np.float, np.float32, np.float64, float]
    return any(obj.dtype == attr for attr in attrs)


def Grading(answer_path, solution_path, csv_file='grading.csv'):
    csv = open(csv_file, 'w')
    cnt = 0
    for fn in os.listdir(answer_path):
        if fn.lower().endswith('.xlsx'):
            student_id = fn[:-5]
            student_id = student_id.replace(" ", "").replace('-', '')
            # print(student_id[0:2]+'-'+student_id[2:8]+'-'+student_id[8:12]+'-'+student_id[-1])
            sol_fn = os.path.join(solution_path, student_id + '.pk')
            if os.path.exists(sol_fn) :
                # print('student_id: ', student_id)
                # print('ans_path: ', answer_path)
                solutions = pickle.load(open(sol_fn, 'rb'))
                try :
                    wb = load_workbook(os.path.join(answer_path, fn))
                    ws = wb.active
                    csv.write((student_id[0:2]+'-'+student_id[2:8]+'-'+student_id[8:12]+'-'+student_id[-1]))
                    # csv.write(student_id)
                    cnt += 1
                    sol_cnt = 0
                    for solution in solutions:
                        score = 0
                        sol_cnt += 1
                        # print(sol_cnt)
                        ans = ws[solution.excel_cells]

                        # print(ws[solution.excel_cells])
                        if isinstance(ans, tuple):
                            ans = np.array([[i.value for i in j] for j in ans])
                        else:
                            ans = np.array([ans.value])
                            if type(ans).__module__ == np.__name__: # ndarray detect
                                if ans[0] is None:
                                    ans = ''
                                else:
                                    ans = ans[0].item()

                        if isinstance(solution.solution, str): # Sentence or Choice
                            score = 0
                            if len(solution.solution) == 1:  # Choice
                                if type(solution.solution) == 'str' :
                                    ans = ans.lower().replace(".", "")
                                else:
                                    ans = ans.replace(".", "")

                                # print(type(ans))
                                # print(type(solution.solution))
                                # ans = ans.lower().replace(".", "")
                                score = np.sum(ans == solution.solution)*solution.score

                            else:  # Sentence
                                score = np.sum(ans == solution.solution) * solution.score
                                    # score = np.sum(
                                    #     ans.astype(np.str) == solution.solution.astype(np.str)) / ans.size * solution.score

                        else:  # Numerical answer
                            # print('ans class ', type(ans))
                            # print('ans ', ans)
                            # print('solution class ', type(solution.solution))
                            # print('solution ', solution.solution)

                            if ans == "":
                                score = 0
                            else:
                                if type(ans) == float or type(ans) == int:
                                    # print('first chk')
                                    if type(solution.solution) != float and type(solution.solution) != int:
                                        score = (np.abs(float(ans) - float(solution.solution.item())))/float(solution.solution.item())
                                        # print(solution.score)
                                        # print(int((score <= percent_acc) == True))
                                        score = int((score <= percent_acc) == True)* solution.score
                                        # print(score)
                                    elif type(solution.solution) == int:
                                        # print('sec chk')
                                        if solution.solution == 0:
                                            divider_sol = 1
                                        else:
                                            divider_sol = solution.solution
                                        score = (np.abs(ans - solution.solution))/divider_sol
                                        score = int((score <= percent_acc) == True)* solution.score
                                    elif type(solution.solution) == float:
                                        # print('thrd chk')
                                        if solution.solution == 0:
                                            divider_sol = 1.0
                                        else:
                                            divider_sol = solution.solution
                                        score = (np.abs(np.abs(ans) - np.abs(solution.solution)))/divider_sol
                                        # print('percent ', score)
                                        score = int((score <= percent_acc) == True)* solution.score

                                elif type(ans) == str:

                                    if type(solution.solution) == int:
                                        # print('str chk')
                                        ans = int(ans)
                                        if solution.solution == 0:
                                            divider_sol = 1
                                        else:
                                            divider_sol = solution.solution
                                        score = (np.abs(ans - solution.solution))/divider_sol
                                        score = int((score <= percent_acc) == True)* solution.score
                                    elif type(solution.solution) == float:
                                        # print('thrd chk')
                                        ans = float(ans)
                                        if solution.solution == 0:
                                            divider_sol = 1.0
                                        else:
                                            divider_sol = solution.solution
                                        score = (np.abs(np.abs(ans) - np.abs(solution.solution)))/divider_sol
                                        # print('percent ', score)
                                        score = int((score <= percent_acc) == True)* solution.score

                                elif is_numeric(ans):
                                    # print('fourth chk')
                                    score = np.abs(ans - solution.solution)
                                    score = np.sum(score <= 0.1) / ans.size * solution.score
                                else:
                                    score = 0

                        # print('score = ', score)
                        csv.write(',' + str(score))
                    csv.write('\n')
                except Exception as error:
                    print('Error : ', error, ' of : ', student_id)
            else:
                print(student_id)
    csv.close()
    print('number of student ', cnt)
